#!/usr/bin/env python3
"""
GreenReach Recipe Bridge v2

Watches the Excel workbook defined in the GR_RECIPES_XLSX environment variable
and publishes normalized plan and schedule payloads to the Charlie server
(typically running on http://127.0.0.1:8091).

Key behaviours:
- Recipes sheet is the source of truth for lighting plans.
- Optional Schedules sheet describes controller schedules.
- Optional Lights sheet lets operators keep device names in sync.
- Additional sheets with daily rows (e.g. "Schedule A", "Herbs") are parsed
  into plan.light.days[] and plan.env.days[].
- Green channel values are evenly split into CW/WW when either white channel is
  missing (preserving previous bridge behaviour).
"""

from __future__ import annotations

import datetime as dt
import json
import logging
import os
import re
import sys
import time
from dataclasses import dataclass
from hashlib import sha256
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

import requests
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

LOGGER = logging.getLogger("recipe-bridge")
LOG_LEVEL = os.environ.get("LOG_LEVEL", "INFO").upper()
logging.basicConfig(
    level=getattr(logging, LOG_LEVEL, logging.INFO),
    format="[%(asctime)s] %(levelname)s %(message)s",
)

DEFAULT_BASE_URL = "http://127.0.0.1:8091"
REQUEST_TIMEOUT = 10
FILE_POLL_SECONDS = float(os.environ.get("GR_BRIDGE_POLL", "2"))

RECIPES_SHEET_NAME = "Recipes"
SCHEDULES_SHEET_NAME = "Schedules"
LIGHTS_SHEET_NAME = "Lights"

HEADER_ALIASES: Dict[str, Tuple[str, ...]] = {
    "name": ("name", "plan", "recipe name", "recipe"),
    "blue": ("blue", "bl", "blue%", "blue pct"),
    "red": ("red", "rd", "red%", "red pct"),
    "cw": ("cw", "cool white", "coolwhite"),
    "ww": ("ww", "warm white", "warmwhite"),
    "green": ("green", "gr"),
    "photoperiod": ("photoperiod", "photoperiod (h)", "photoperiod h", "photoperiod hours"),
    "sunrise": ("sunrise", "sunrise (min)", "sunrise min", "ramp up", "rampup", "ramp up min"),
    "sunset": ("sunset", "sunset (min)", "sunset min", "ramp down", "rampdown", "ramp down min"),
    "day": ("day", "d"),
    "stage": ("stage",),
    "ppfd": ("ppfd", "ppfd (umol)", "ppfd umol"),
    "photoperiod_row": ("photoperiod", "photoperiod h", "photoperiod (h)", "light photoperiod"),
    "temp_c": ("temp", "temp c", "tempc", "temperature", "temperature c", "temperature (c)"),
    "rh": ("rh", "humidity", "relative humidity"),
    "rh_band": ("rh band", "rh ±", "rh +/-", "rh range"),
    "co2": ("co2", "co2 ppm"),
    "id": ("id", "fixture id", "light id", "device id"),
    "start": ("start", "start time"),
    "duration_hours": ("duration hours", "duration", "hours", "duration (h)"),
    "ramp_up": ("ramp up", "rampup", "ramp up min", "rampupmin"),
    "ramp_down": ("ramp down", "rampdown", "ramp down min", "rampdownmin"),
    "override_mode": ("override", "override mode"),
    "override_val": ("override val", "override value"),
    "recipe_ref": ("recipe", "plan", "plan key", "plan id"),
    "seed_date": ("seed date", "seeddate", "seed", "seeded"),
}


@dataclass
class WorkbookPayload:
    plans: List[Dict[str, Any]]
    schedules: List[Dict[str, Any]]
    lights: List[Dict[str, Any]]


def normalize_header(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip().lower()
    text = text.replace("%", " percent ")
    return re.sub(r"[^a-z0-9]+", " ", text).strip().replace(" ", "_")


def build_header_map(sheet: Worksheet) -> Dict[str, int]:
    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header_row:
        return {}
    mapping: Dict[str, int] = {}
    for idx, cell in enumerate(header_row):
        key = normalize_header(cell)
        if key:
            mapping[key] = idx
    return mapping


def find_column(header_map: Dict[str, int], key: str) -> Optional[int]:
    candidates: Iterable[str] = (key,) + HEADER_ALIASES.get(key, tuple())
    for candidate in candidates:
        normalized = normalize_header(candidate)
        if normalized in header_map:
            return header_map[normalized]
    return None


def coerce_number(value: Any) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        num = float(value)
        if not (num != num or num in (float("inf"), float("-inf"))):  # NaN/inf guard
            return num
        return None
    if isinstance(value, str):
        cleaned = value.strip()
        if not cleaned:
            return None
        cleaned = cleaned.replace(",", "")
        cleaned = cleaned.replace("%", "")
        match = re.search(r"-?\d+(?:\.\d+)?", cleaned)
        if not match:
            return None
        try:
            return float(match.group(0))
        except ValueError:
            return None
    return None


def coerce_int(value: Any) -> Optional[int]:
    num = coerce_number(value)
    if num is None:
        return None
    try:
        return int(round(num))
    except (TypeError, ValueError):
        return None


def coerce_str(value: Any) -> Optional[str]:
    if value is None:
        return None
    if isinstance(value, str):
        text = value.strip()
        return text or None
    return str(value)


def split_green_into_whites(cw: Optional[float], ww: Optional[float], green: Optional[float]) -> Tuple[Optional[float], Optional[float]]:
    """
    Split green percentage into CW and WW using SPD-weighted method.
    
    SPD-Weighted Split (default):
    - CW contributes ~35% of green output (500-600nm)
    - WW contributes ~65% of green output (500-600nm)
    - Based on typical phosphor-converted white LED spectra
    
    Fallback to 50/50 if either CW or WW is already specified.
    """
    if green is None:
        return cw, ww
    
    cw_val = cw if (cw is not None and cw > 0) else None
    ww_val = ww if (ww is not None and ww > 0) else None
    
    if cw_val is None and ww_val is None:
        # SPD-weighted split based on typical cool/warm white LED spectra
        # CW (5000-6500K): ~35% green contribution
        # WW (2700-3500K): ~65% green contribution
        # Derived from integrate(BASIS.cw, 500, 600) vs integrate(BASIS.ww, 500, 600)
        cw_ratio = 0.35
        ww_ratio = 0.65
        return green * cw_ratio, green * ww_ratio
    
    if cw_val is None:
        # WW specified, split remaining green 50/50 for safety
        return green / 2.0, ww_val
    
    if ww_val is None:
        # CW specified, split remaining green 50/50 for safety
        return cw_val, green / 2.0
    
    # Both specified, return as-is
    return cw_val, ww_val


def clamp_percent(value: Optional[float]) -> Optional[float]:
    if value is None:
        return None
    clipped = max(0.0, min(100.0, float(value)))
    return round(clipped, 4)


def clamp_minutes(value: Optional[int]) -> Optional[int]:
    if value is None:
        return None
    return max(0, min(120, int(value)))


def parse_duration_hours(value: Any) -> Optional[int]:
    if value is None:
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return max(0, min(24, int(round(float(value)))))
    text = coerce_str(value)
    if not text:
        return None
    candidate = text.replace("hours", " ").replace("hrs", " ")
    if "/" in candidate:
        candidate = candidate.split("/", 1)[0]
    match = re.search(r"-?\d+(?:\.\d+)?", candidate)
    if not match:
        return None
    try:
        number = float(match.group(0))
    except ValueError:
        return None
    return max(0, min(24, int(round(number))))


def parse_time_of_day(value: Any) -> Optional[str]:
    if value is None:
        return None
    if isinstance(value, dt.datetime):
        return parse_time_of_day(value.time())
    if isinstance(value, dt.time):
        base = value.replace(second=0, microsecond=0)
        return base.strftime("%H:%M")
    text = coerce_str(value)
    if not text:
        return None
    normalized = re.sub(r"\s+", " ", text.strip())
    upper = normalized.upper()
    formats = ["%H:%M", "%H%M", "%I:%M%p", "%I:%M %p", "%I%p", "%I %p"]
    for fmt in formats:
        try:
            parsed = dt.datetime.strptime(upper, fmt)
            return parsed.time().strftime("%H:%M")
        except ValueError:
            continue
    digits = re.sub(r"[^0-9]", "", normalized)
    if digits:
        if len(digits) == 3:
            digits = f"0{digits}"
        if len(digits) == 4:
            try:
                parsed = dt.datetime.strptime(digits, "%H%M")
                return parsed.time().strftime("%H:%M")
            except ValueError:
                pass
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        hours = int(value)
        minutes = int(round((float(value) - hours) * 60))
        while minutes >= 60:
            minutes -= 60
            hours += 1
        hours %= 24
        return f"{hours:02d}:{minutes:02d}"
    return None


def normalize_seed_date(value: Any) -> Optional[str]:
    if value is None:
        return None
    if isinstance(value, dt.datetime):
        return value.date().isoformat()
    if isinstance(value, dt.date):
        return value.isoformat()
    text = coerce_str(value)
    if not text:
        return None
    normalized = text.strip()
    candidates = [
        "%Y-%m-%d",
        "%Y/%m/%d",
        "%m/%d/%Y",
        "%m/%d/%y",
        "%d/%m/%Y",
        "%d/%m/%y",
        "%d-%m-%Y",
        "%d-%m-%y",
    ]
    for fmt in candidates:
        try:
            parsed = dt.datetime.strptime(normalized, fmt)
            return parsed.date().isoformat()
        except ValueError:
            continue
    digits = re.sub(r"[^0-9]", "", normalized)
    if len(digits) == 8:
        try:
            parsed = dt.datetime.strptime(digits, "%Y%m%d")
            return parsed.date().isoformat()
        except ValueError:
            pass
    if len(digits) == 6:
        try:
            parsed = dt.datetime.strptime(digits, "%y%m%d")
            return parsed.date().isoformat()
        except ValueError:
            pass
    return None


def make_plan_id(name: str) -> str:
    slug = re.sub(r"[^A-Za-z0-9]+", "-", name.strip()).strip("-")
    return slug or "Plan"


def parse_photoperiod(value: Any) -> Optional[int]:
    return parse_duration_hours(value)


def parse_recipes_sheet(sheet: Worksheet) -> Tuple[List[Dict[str, Any]], Dict[str, Dict[str, Any]]]:
    header_map = build_header_map(sheet)
    name_col = find_column(header_map, "name")
    if name_col is None:
        raise ValueError("Recipes sheet must include a Name column.")
    blue_col = find_column(header_map, "blue")
    red_col = find_column(header_map, "red")
    cw_col = find_column(header_map, "cw")
    ww_col = find_column(header_map, "ww")
    green_col = find_column(header_map, "green")
    photoperiod_col = find_column(header_map, "photoperiod")
    sunrise_col = find_column(header_map, "sunrise")
    sunset_col = find_column(header_map, "sunset")

    plans: List[Dict[str, Any]] = []
    plan_lookup: Dict[str, Dict[str, Any]] = {}

    for row in sheet.iter_rows(min_row=2, values_only=True):
        raw_name = row[name_col] if name_col is not None else None
        name = coerce_str(raw_name)
        if not name:
            continue
        blue = clamp_percent(coerce_number(row[blue_col]) if blue_col is not None else None)
        red = clamp_percent(coerce_number(row[red_col]) if red_col is not None else None)
        cw_raw = coerce_number(row[cw_col]) if cw_col is not None else None
        ww_raw = coerce_number(row[ww_col]) if ww_col is not None else None
        green = coerce_number(row[green_col]) if green_col is not None else None
        cw_split, ww_split = split_green_into_whites(cw_raw, ww_raw, green)
        cw = clamp_percent(cw_split)
        ww = clamp_percent(ww_split)
        photoperiod = parse_photoperiod(row[photoperiod_col]) if photoperiod_col is not None else None
        sunrise = clamp_minutes(coerce_int(row[sunrise_col]) if sunrise_col is not None else None)
        sunset = clamp_minutes(coerce_int(row[sunset_col]) if sunset_col is not None else None)

        plan_id = make_plan_id(name)
        plan = {
            "id": plan_id,
            "key": plan_id,
            "name": name,
        }
        if photoperiod is not None:
            plan["photoperiod"] = photoperiod
        ramp: Dict[str, int] = {}
        if sunrise is not None:
            ramp["sunrise"] = sunrise
        if sunset is not None:
            ramp["sunset"] = sunset
        if ramp:
            plan["ramp"] = ramp

        day_entry: Dict[str, Any] = {"stage": "Static"}
        if blue is not None:
            day_entry["bl"] = blue
        if red is not None:
            day_entry["rd"] = red
        if cw is not None:
            day_entry["cw"] = cw
        if ww is not None:
            day_entry["ww"] = ww
        plan["days"] = [day_entry]
        plan["meta"] = {"source": "excel"}

        plans.append(plan)
        plan_lookup[make_lookup_key(name)] = plan
        plan_lookup[make_lookup_key(plan_id)] = plan

    return plans, plan_lookup


def make_lookup_key(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", value.lower()) if value else ""


def parse_lights_sheet(sheet: Worksheet, plan_lookup: Dict[str, Dict[str, Any]]) -> List[Dict[str, Any]]:
    header_map = build_header_map(sheet)
    id_col = find_column(header_map, "id")
    name_col = find_column(header_map, "name")
    recipe_col = find_column(header_map, "recipe_ref")
    if id_col is None:
        LOGGER.warning("Lights sheet present but missing an ID column. Skipping.")
        return []

    lights: List[Dict[str, Any]] = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        identifier = coerce_str(row[id_col]) if id_col is not None else None
        if not identifier:
            continue
        name = coerce_str(row[name_col]) if name_col is not None else None
        recipe_name = coerce_str(row[recipe_col]) if recipe_col is not None else None
        plan = plan_lookup.get(make_lookup_key(recipe_name)) if recipe_name else None
        plan_id = plan.get("id") if plan else None
        lights.append({
            "id": identifier,
            "name": name,
            "recipe": plan_id or recipe_name,
        })
    return lights


def parse_schedules_sheet(sheet: Worksheet, plan_lookup: Dict[str, Dict[str, Any]]) -> List[Dict[str, Any]]:
    header_map = build_header_map(sheet)
    id_col = find_column(header_map, "id")
    if id_col is None:
        LOGGER.warning("Schedules sheet missing ID column. Skipping schedule sync.")
        return []

    start_col = find_column(header_map, "start")
    duration_col = find_column(header_map, "duration_hours")
    ramp_up_col = find_column(header_map, "ramp_up")
    ramp_down_col = find_column(header_map, "ramp_down")
    override_mode_col = find_column(header_map, "override_mode")
    override_val_col = find_column(header_map, "override_val")
    recipe_col = find_column(header_map, "recipe_ref")
    seed_col = find_column(header_map, "seed_date")

    schedules: List[Dict[str, Any]] = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        identifier = coerce_str(row[id_col]) if id_col is not None else None
        if not identifier:
            continue
        schedule: Dict[str, Any] = {"id": identifier}
        if start_col is not None:
            start_value = parse_time_of_day(row[start_col])
            if start_value:
                schedule["start"] = start_value
        if duration_col is not None:
            duration = parse_duration_hours(row[duration_col])
            if duration is not None:
                schedule["durationHours"] = duration
        if ramp_up_col is not None:
            value = clamp_minutes(coerce_int(row[ramp_up_col]))
            if value is not None:
                schedule["rampUpMin"] = value
        if ramp_down_col is not None:
            value = clamp_minutes(coerce_int(row[ramp_down_col]))
            if value is not None:
                schedule["rampDownMin"] = value
        if override_mode_col is not None:
            mode = coerce_str(row[override_mode_col])
            if mode:
                schedule["overrideMode"] = mode.lower()
        if override_val_col is not None:
            value = row[override_val_col]
            str_value = coerce_str(value)
            num_value = coerce_number(value)
            if str_value and num_value is None:
                schedule["overrideVal"] = str_value
            elif num_value is not None:
                schedule["overrideVal"] = num_value
        if recipe_col is not None:
            recipe_name = coerce_str(row[recipe_col])
            if recipe_name:
                plan = plan_lookup.get(make_lookup_key(recipe_name))
                if plan:
                    schedule["planKey"] = plan["id"]
                else:
                    schedule["planKey"] = recipe_name
        if seed_col is not None:
            seed = normalize_seed_date(row[seed_col])
            if seed:
                schedule["seedDate"] = seed
        schedules.append(schedule)
    return schedules


def parse_daily_sheet(sheet: Worksheet) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
    header_map = build_header_map(sheet)
    if not header_map:
        return [], []
    day_col = find_column(header_map, "day")
    stage_col = find_column(header_map, "stage")
    ppfd_col = find_column(header_map, "ppfd")
    photoperiod_col = find_column(header_map, "photoperiod_row")
    blue_col = find_column(header_map, "blue")
    red_col = find_column(header_map, "red")
    cw_col = find_column(header_map, "cw")
    ww_col = find_column(header_map, "ww")
    green_col = find_column(header_map, "green")
    temp_col = find_column(header_map, "temp_c")
    rh_col = find_column(header_map, "rh")
    rh_band_col = find_column(header_map, "rh_band")
    co2_col = find_column(header_map, "co2")

    light_days: List[Dict[str, Any]] = []
    env_days: List[Dict[str, Any]] = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        day_value = coerce_int(row[day_col]) if day_col is not None else None
        stage = coerce_str(row[stage_col]) if stage_col is not None else None
        ppfd = coerce_number(row[ppfd_col]) if ppfd_col is not None else None
        photoperiod = parse_photoperiod(row[photoperiod_col]) if photoperiod_col is not None else None
        blue = clamp_percent(coerce_number(row[blue_col]) if blue_col is not None else None)
        red = clamp_percent(coerce_number(row[red_col]) if red_col is not None else None)
        cw_raw = coerce_number(row[cw_col]) if cw_col is not None else None
        ww_raw = coerce_number(row[ww_col]) if ww_col is not None else None
        green = coerce_number(row[green_col]) if green_col is not None else None
        cw_split, ww_split = split_green_into_whites(cw_raw, ww_raw, green)
        cw = clamp_percent(cw_split)
        ww = clamp_percent(ww_split)

        temp_c = coerce_number(row[temp_col]) if temp_col is not None else None
        rh = coerce_number(row[rh_col]) if rh_col is not None else None
        rh_band = coerce_number(row[rh_band_col]) if rh_band_col is not None else None
        co2 = coerce_number(row[co2_col]) if co2_col is not None else None

        has_light = any(value is not None for value in (blue, red, cw, ww, ppfd, photoperiod, stage))
        has_env = any(value is not None for value in (temp_c, rh, rh_band, co2))
        if not has_light and not has_env:
            continue

        day_index = day_value
        if day_index is None:
            day_index = len(light_days) + 1 if has_light else len(env_days) + 1

        if has_light:
            entry: Dict[str, Any] = {"d": day_index}
            if stage:
                entry["stage"] = stage
            if ppfd is not None:
                entry["ppfd"] = ppfd
            if photoperiod is not None:
                entry["photoperiod"] = photoperiod
            mix: Dict[str, float] = {}
            if blue is not None:
                mix["bl"] = blue
            if red is not None:
                mix["rd"] = red
            if cw is not None:
                mix["cw"] = cw
            if ww is not None:
                mix["ww"] = ww
            if mix:
                entry["mix"] = mix
            light_days.append(entry)

        if has_env:
            env_entry: Dict[str, Any] = {"d": day_index}
            if temp_c is not None:
                env_entry["tempC"] = temp_c
            if rh is not None:
                clipped_rh = clamp_percent(rh)
                if clipped_rh is not None:
                    env_entry["rh"] = clipped_rh
            if rh_band is not None:
                band = clamp_percent(abs(rh_band))
                if band is not None:
                    env_entry["rhBand"] = band
            if co2 is not None:
                env_entry["co2"] = co2
            env_days.append(env_entry)

    light_days.sort(key=lambda item: item.get("d", 0))
    env_days.sort(key=lambda item: item.get("d", 0))
    return light_days, env_days


def merge_daily_sheets(workbook, plan_lookup: Dict[str, Dict[str, Any]]) -> None:
    for sheet_name in workbook.sheetnames:
        if sheet_name in {RECIPES_SHEET_NAME, SCHEDULES_SHEET_NAME, LIGHTS_SHEET_NAME}:
            continue
        sheet = workbook[sheet_name]
        light_days, env_days = parse_daily_sheet(sheet)
        if not light_days and not env_days:
            continue
        key = make_lookup_key(sheet_name)
        plan = plan_lookup.get(key)
        if plan is None:
            plan_id = make_plan_id(sheet_name)
            LOGGER.info("Creating plan shell for daily sheet '%s' (id=%s)", sheet_name, plan_id)
            plan = {
                "id": plan_id,
                "key": plan_id,
                "name": sheet_name,
                "days": [],
                "meta": {"source": "excel"},
            }
            plan_lookup[key] = plan
            plan_lookup[make_lookup_key(plan_id)] = plan
            plan_lookup[make_lookup_key(plan["name"])] = plan
        plan.setdefault("kind", "cropPlan")
        defaults: Dict[str, Any] = dict(plan.get("defaults") or {})
        if "photoperiod" in plan and "photoperiod" not in defaults:
            defaults["photoperiod"] = plan["photoperiod"]
        if "ramp" in plan and "ramp" not in defaults:
            defaults["ramp"] = plan["ramp"]
        if defaults:
            plan["defaults"] = defaults
        if light_days:
            plan.setdefault("light", {})["days"] = light_days
        if env_days:
            plan.setdefault("env", {})["days"] = env_days


def load_workbook_payload(path: str) -> WorkbookPayload:
    LOGGER.info("Reading workbook: %s", path)
    workbook = load_workbook(path, data_only=True, read_only=False)
    if RECIPES_SHEET_NAME not in workbook.sheetnames:
        raise ValueError(f"Workbook missing required sheet '{RECIPES_SHEET_NAME}'.")

    recipes_sheet = workbook[RECIPES_SHEET_NAME]
    plans, plan_lookup = parse_recipes_sheet(recipes_sheet)
    merge_daily_sheets(workbook, plan_lookup)

    lights: List[Dict[str, Any]] = []
    if LIGHTS_SHEET_NAME in workbook.sheetnames:
        lights = parse_lights_sheet(workbook[LIGHTS_SHEET_NAME], plan_lookup)

    schedules: List[Dict[str, Any]] = []
    if SCHEDULES_SHEET_NAME in workbook.sheetnames:
        schedules = parse_schedules_sheet(workbook[SCHEDULES_SHEET_NAME], plan_lookup)

    unique_plans = {plan["id"]: plan for plan in plan_lookup.values() if plan.get("id")}
    ordered_plans = sorted(unique_plans.values(), key=lambda item: item["name"].lower())
    return WorkbookPayload(plans=ordered_plans, schedules=schedules, lights=lights)


def bool_env(name: str, default: bool = False) -> bool:
    raw = os.environ.get(name)
    if raw is None:
        return default
    return raw.strip().lower() in {"1", "true", "yes", "on"}


def get_headers() -> Dict[str, str]:
    headers = {"Content-Type": "application/json"}
    pin = os.environ.get("FARM_PIN") or os.environ.get("CTRL_PIN")
    if pin:
        headers["x-farm-pin"] = pin
    return headers


def post_json(session: requests.Session, base_url: str, path: str, payload: Any) -> Optional[requests.Response]:
    url = f"{base_url.rstrip('/')}/{path.lstrip('/')}"
    headers = get_headers()
    try:
        response = session.post(url, json=payload, headers=headers, timeout=REQUEST_TIMEOUT)
        if response.ok:
            LOGGER.info("POST %s → %s", path, response.status_code)
        else:
            LOGGER.error("POST %s failed (%s): %s", path, response.status_code, response.text)
        return response
    except requests.RequestException as exc:
        LOGGER.error("POST %s failed: %s", path, exc)
        return None


def patch_device_name(session: requests.Session, base_url: str, device_id: str, name: str) -> None:
    url = f"{base_url.rstrip('/')}/devices/{device_id}"
    headers = get_headers()
    payload = {"id": device_id, "name": name}
    try:
        response = session.patch(url, json=payload, headers=headers, timeout=REQUEST_TIMEOUT)
        if response.ok:
            LOGGER.info("PATCH /devices/%s → %s", device_id, response.status_code)
        else:
            LOGGER.warning("Failed to patch /devices/%s (%s): %s", device_id, response.status_code, response.text)
    except requests.RequestException as exc:
        LOGGER.warning("Device patch failed for %s: %s", device_id, exc)


def sync_lights(session: requests.Session, base_url: str, lights: Sequence[Dict[str, Any]], enable_names: bool) -> None:
    if not lights:
        return
    for light in lights:
        identifier = light.get("id")
        if not identifier:
            continue
        name = light.get("name")
        if enable_names and name:
            patch_device_name(session, base_url, identifier, name)


def payload_digest(payload: WorkbookPayload) -> str:
    serializable = {
        "plans": payload.plans,
        "schedules": payload.schedules,
        "lights": payload.lights,
    }
    data = json.dumps(serializable, sort_keys=True, separators=(",", ":"))
    return sha256(data.encode("utf-8")).hexdigest()


def run_bridge_once(path: str) -> WorkbookPayload:
    payload = load_workbook_payload(path)
    LOGGER.info(
        "Parsed %d plan(s), %d schedule(s), %d light(s)",
        len(payload.plans),
        len(payload.schedules),
        len(payload.lights),
    )
    return payload


def watch_loop(path: str, base_url: str, sync_names: bool, sync_schedule: bool) -> None:
    session = requests.Session()
    last_digest: Optional[str] = None
    last_mtime: Optional[float] = None
    LOGGER.info("Watching %s for changes (plans → %s)", path, base_url)
    while True:
        try:
            stat = os.stat(path)
        except FileNotFoundError:
            LOGGER.warning("Workbook not found at %s. Waiting...", path)
            time.sleep(FILE_POLL_SECONDS)
            continue
        mtime = stat.st_mtime
        if last_mtime is not None and mtime <= last_mtime:
            time.sleep(FILE_POLL_SECONDS)
            continue
        last_mtime = mtime
        # Allow brief settle time for in-progress writes
        time.sleep(0.5)
        try:
            payload = run_bridge_once(path)
        except Exception as exc:  # pylint: disable=broad-except
            LOGGER.exception("Failed to parse workbook: %s", exc)
            time.sleep(FILE_POLL_SECONDS)
            continue
        digest = payload_digest(payload)
        if digest == last_digest:
            LOGGER.debug("Workbook changed but produced identical payload. Skipping POST.")
            time.sleep(FILE_POLL_SECONDS)
            continue
        last_digest = digest

        if payload.plans:
            post_json(session, base_url, "/plans", {"plans": payload.plans})
        else:
            LOGGER.warning("No plans parsed from workbook; skipping /plans POST.")

        if sync_schedule:
            post_json(session, base_url, "/sched", {"schedules": payload.schedules})
        elif payload.schedules:
            LOGGER.info("Schedules available but GR_SYNC_SCHEDULE disabled; skipping /sched POST.")

        sync_lights(session, base_url, payload.lights, enable_names=sync_names)

        time.sleep(FILE_POLL_SECONDS)


def main(argv: Sequence[str]) -> int:
    workbook_path = os.environ.get("GR_RECIPES_XLSX")
    if not workbook_path:
        LOGGER.error("GR_RECIPES_XLSX environment variable is required.")
        return 1
    base_url = os.environ.get("PROXY_BASE", DEFAULT_BASE_URL)
    sync_names = bool_env("GR_SYNC_NAMES", default=True)
    sync_schedule = bool_env("GR_SYNC_SCHEDULE", default=True)
    LOGGER.info(
        "Recipe Bridge configured (base=%s, sync_names=%s, sync_schedule=%s)",
        base_url,
        sync_names,
        sync_schedule,
    )
    try:
        watch_loop(workbook_path, base_url, sync_names=sync_names, sync_schedule=sync_schedule)
    except KeyboardInterrupt:
        LOGGER.info("Bridge stopped via keyboard interrupt.")
        return 0
    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv))
